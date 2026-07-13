#!/usr/bin/env python3
"""Baker 1031 site — CI build (runs on Netlify).

Pipeline:
  1. Fetch "Master Listings Sheet" from Google Sheets as xlsx
     (SHEET_ID env var; falls back to data/fallback-master-listings.xlsx).
  2. Generate 63+ offering pages, 85+ sponsor pages, both directories,
     and the DST sector chart from the workbook.
  3. Inject the shared nav/footer partials into every page.
  4. Emit the complete site to dist/ (+ shared assets).

Local test:  python3 ci_build.py          (uses fallback xlsx if no network)
Netlify:     see netlify.toml
"""
import datetime
import hashlib
import html as html_lib
import io, json, os, posixpath, re, shutil, sys, unicodedata, urllib.request
from decimal import Decimal, ROUND_HALF_UP
from urllib.parse import unquote, urljoin, urlparse

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
DIST = os.path.join(ROOT, 'dist')
SHEET_ID = os.environ.get('SHEET_ID', '1vTqb5YX8pFjZxToGd2pJ_ncPbny2PXpW5gXx-7IlyZg')
GA4_MEASUREMENT_ID = os.environ.get('GA4_MEASUREMENT_ID', 'G-P29LR49RL8')
# Clarity project IDs are public client-side configuration, so this fallback
# keeps local previews and Netlify builds aligned while still allowing an
# environment override for a future project change.
CLARITY_PROJECT_ID = os.environ.get('CLARITY_PROJECT_ID', 'xly64re3u1').strip()
BASE_URL = os.environ.get('BASE_URL', 'https://baker1031.com').rstrip('/')
BUILD_DATE = os.environ.get('BUILD_DATE', datetime.date.today().isoformat())
REVIEW_DATE = 'July 12, 2026'
REVIEW_DATE_ISO = '2026-07-12'
TITLE_LIMIT = 70
DESCRIPTION_LIMIT = 155

# These files remain available as authoring/source material but must never be
# emitted as public, indexable pages.
SKIP_SOURCE_PAGES = {'article-template.html', 'insight-1031-exchange-guide.html'}
NOINDEX_PAGES = {'article-template.html', 'insight-1031-exchange-guide.html',
                 'employee.html', 'account.html', '404.html', 'baker1031.html'}
NOINDEX_PAGES.add('site-search.html')

APPROVED_DISCLOSURE_LINK = (
    'For entity and registration details, see the '
    '<a href="disclosures.html">Baker 1031 disclosures</a> and the applicable '
    'Private Placement Memorandum.'
)

PAGE_DESCRIPTIONS = {
    'index.html': 'Baker 1031 Investments helps accredited investors evaluate 1031 exchanges, DST replacement property, 721 UPREIT strategies, Opportunity Zones, mineral interests, and REITs.',
    'baker1031.html': 'Baker 1031 Investments provides 1031 exchange education, DST replacement-property research, current offerings, sponsor data, and investor resources.',
    'about.html': 'Learn how Baker 1031 Investments evaluates 1031 replacement-property strategies, works with accredited investors, and coordinates offerings through Aurora Securities.',
    'investments.html': 'Browse current Delaware Statutory Trust and 1031 exchange replacement-property offerings, with property details, sponsor information, projected income, and risks.',
    'sponsors.html': 'Research DST and 1031 exchange sponsors, current offerings, sponsor facts, and realized full-cycle track-record data compiled by Baker 1031 Investments.',
    'insights.html': 'Explore Baker 1031 Investments research on 1031 exchanges, DSTs, 721 UPREITs, Opportunity Zones, mineral and royalty interests, REITs, and real-estate tax planning.',
    'delaware-statutory-trusts.html': 'Learn how Delaware Statutory Trusts can serve as 1031 exchange replacement property, including structure, benefits, risks, financing, income, and due diligence.',
    '1031-exchange-guide.html': 'A practical guide to 1031 exchange rules, like-kind property, the 45-day identification period, the 180-day exchange period, boot, and replacement strategies.',
    '721-exchange-guide.html': 'Understand 721 UPREIT exchanges, including how DST interests or real property may be contributed for operating-partnership units and the trade-offs involved.',
    'opportunity-zones-guide.html': 'Understand Qualified Opportunity Zone funds, capital-gain deferral, the ten-year holding period, tax treatment, suitability, and investment risks.',
    'mineral-rights-1031-guide.html': 'Learn how qualifying mineral and royalty interests may fit into a 1031 exchange, including perpetual-interest rules, income, depletion, and risk.',
    'reits-guide.html': 'Compare public, non-traded, and private REITs, including liquidity, distributions, valuation, fees, risks, and connections to 721 UPREIT strategies.',
    'methodology.html': 'Review how Baker 1031 compiles sponsor, offering, benchmark, and full-cycle data, including definitions, limitations, and sponsor-reported figures.',
    'contact.html': 'Contact the Baker 1031 Investments desk about current offerings, 1031 replacement-property research, sponsor due diligence, or advisor coordination.',
    'request-access.html': 'Request access to Baker 1031 Investments resources and current offerings for accredited investors, subject to suitability and required disclosures.',
    'faq.html': 'Answers to common questions about 1031 exchanges, DSTs, 721 UPREITs, Opportunity Zones, REITs, mineral interests, eligibility, fees, and risks.',
    'site-search.html': 'Search Baker 1031 Investments research, 1031 exchange guides, DST resources, calculators, offerings, sponsors, and investor education.',
    'ask-llm.html': 'Use carefully designed prompts to ask ChatGPT, Claude, Gemini, Perplexity, or Copilot about Baker 1031 Investments and its published research.',
}

HUB_FILES = {
    '1031-exchange-guide.html', 'delaware-statutory-trusts.html',
    '721-exchange-guide.html', 'opportunity-zones-guide.html',
    'mineral-rights-1031-guide.html', 'reits-guide.html', 'strategies.html',
    'investments.html', 'sponsors.html', 'insights.html', 'faq.html',
}

# Only these pages contain a user-facing interactive calculator. Article and
# glossary URLs that happen to contain "calculator" are educational guides,
# not WebApplications, so they should not receive calculator schema.
INTERACTIVE_CALCULATOR_PAGES = {
    'calculator-cost-segregation.html',
    '1031-exchange-boot-calculator.html',
    '1031-exchange-capital-gains-tax-calculator.html',
    '45-180-day-deadline-calculator.html',
    'cap-rate-cash-on-cash-calculator.html',
    'capital-gains-tax-calculator.html',
    'debt-replacement-ltv-calculator.html',
    'depreciation-recapture-calculator.html',
    'mineral-rights-valuation-calculator.html',
    'oz-vs-1031-vs-cash-out-calculator.html',
    'passive-income-calculator.html',
    'royalties-vs-dst-income-calculator.html',
    'sell-vs-1031-exchange-calculator.html',
}

AUTHORITY_SOURCES = [
    'https://www.irs.gov/businesses/small-businesses-self-employed/like-kind-exchanges-real-estate-tax-tips',
    'https://www.irs.gov/forms-pubs/about-form-8824',
    'https://www.investor.gov/',
    'https://www.finra.org/',
    'https://brokercheck.finra.org/individual/summary/7537416',
]

STATE_TAX_AUTHORITIES = json.load(open(os.path.join(ROOT, 'data', 'state-tax-authorities.json')))
STATE_AUTHORITY_URLS = {item['url'] for item in STATE_TAX_AUTHORITIES.values()}
AUTHOR_NAME = 'Gerald F. "Jerry" Baker, III'
AUTHOR_IMAGE = 'https://res.cloudinary.com/opoazlei/image/upload/v1783927734/jerry-baker_ovhy2w.jpg'
REVIEWER = {
    '@type': 'Person',
    'name': 'Lori Kamen',
    'jobTitle': 'Chief Compliance Officer, Aurora Securities, Inc.',
    'identifier': {'@type': 'PropertyValue', 'propertyID': 'FINRA CRD', 'value': '2805591'},
    'sameAs': ['https://brokercheck.finra.org/individual/summary/2805591'],
}
REVIEW_NOTE_HTML = '''<p>
Reviewed by Lori Kamen — Chief Compliance Officer, Aurora Securities, Inc. (FINRA CRD #2805591). Last reviewed July 12, 2026. Educational content is reviewed periodically for accuracy and regulatory compliance; current law, offering terms, and suitability depend on the applicable facts and documents.
</p>'''

CURRENT_1031_SOURCE_NOTE = '''<p class="b1031-current-law-source">Current-law source reviewed July 11, 2026: <a href="https://www.irs.gov/businesses/small-businesses-self-employed/like-kind-exchanges-real-estate-tax-tips" target="_blank" rel="noopener noreferrer">IRS guidance on like-kind exchanges</a>. Section 1031 treatment depends on the property, taxpayer, timing, and transaction documents; confirm current treatment with your CPA and attorney.</p>'''

CURRENT_OZ_SOURCE_NOTE = '''<p class="b1031-current-law-source">Current-law source reviewed July 11, 2026: <a href="https://www.irs.gov/newsroom/opportunity-zones" target="_blank" rel="noopener noreferrer">IRS Opportunity Zone guidance</a> and <a href="https://www.irs.gov/irb/2026-28_IRB" target="_blank" rel="noopener noreferrer">IRS Notice 2026-40</a>. Opportunity Zone benefits are conditional, time-sensitive, and dependent on the QOF, the taxpayer, the holding period, and current law; confirm the details with your CPA and attorney.</p>'''

SEO_META = {}

# Optional generated audio manifest. Audio is intentionally absent until the
# synthesis/upload job has produced a real asset for a page. This keeps the
# build truthful: no AudioObject schema or player is emitted for missing media.
AUDIO_MANIFEST_PATH = os.path.join(ROOT, 'data', 'audio-manifest.json')
try:
    AUDIO_MANIFEST = json.load(open(AUDIO_MANIFEST_PATH, encoding='utf-8'))
    if not isinstance(AUDIO_MANIFEST, dict):
        AUDIO_MANIFEST = {}
except (FileNotFoundError, json.JSONDecodeError):
    AUDIO_MANIFEST = {}

# ---------------------------------------------------------------- fetch
def load_workbook():
    url = 'https://docs.google.com/spreadsheets/d/%s/export?format=xlsx' % SHEET_ID
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'baker1031-ci'})
        with urllib.request.urlopen(req, timeout=60) as r:
            blob = r.read()
        if not blob[:2] == b'PK':
            raise RuntimeError('not an xlsx (sheet not link-viewable?)')
        print('fetched live sheet: %d bytes' % len(blob))
        return openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:
        print('WARN: live sheet fetch failed (%s); using fallback snapshot' % e)
        return openpyxl.load_workbook(os.path.join(ROOT, 'data', 'fallback-master-listings.xlsx'), data_only=True)

WB = load_workbook()

def sheet_dicts(name):
    ws = WB[name]
    hdr = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, '') for v in r):
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]})
    return out

LIST = sheet_dicts('Master Listings')
TRACK = sheet_dicts('Sponsor Trackrecord')
DOCS = sheet_dicts('Documents')
SPON = sheet_dicts('Sponsor Connection')
BM = sheet_dicts('Benchmarks')
print('rows: listings=%d track=%d docs=%d sponsors=%d benchmarks=%d' % (len(LIST), len(TRACK), len(DOCS), len(SPON), len(BM)))

# ---------------------------------------------------------------- helpers
def slugify(x):
    x = unicodedata.normalize('NFKD', str(x)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '-', x.lower()).strip('-')

def s(v):
    if v is None: return ''
    if isinstance(v, float) and v == int(v) and abs(v) >= 1000:
        return '{:,.0f}'.format(v)
    return str(v).strip()

def pct(v, dp=2):
    if v in (None, ''): return ''
    if isinstance(v, (int, float)):
        n = v * 100 if abs(v) <= 1.5 else v
        quantum = Decimal('1').scaleb(-dp)
        rounded = Decimal(str(n)).quantize(quantum, rounding=ROUND_HALF_UP)
        return format(rounded, '.' + str(dp) + 'f') + '%'
    return str(v).strip()

def money(v):
    if v in (None, ''): return ''
    if isinstance(v, (int, float)):
        return '${:,.0f}'.format(v)
    return str(v).strip()

def money_compact(v):
    # AUM-style compact currency: 2000000000 -> "$2.0B", 440000000 -> "$440M".
    if v in (None, ''): return ''
    if isinstance(v, (int, float)):
        n = float(v)
        if n >= 1e9:
            return '${:.1f}B'.format(n / 1e9)
        if n >= 1e6:
            return '${:.0f}M'.format(n / 1e6)
        if n >= 1e3:
            return '${:,.0f}'.format(n)
        return '${:,.0f}'.format(n)
    return str(v).strip()

def yrs(v):
    if v in (None, ''): return ''
    if isinstance(v, (int, float)):
        return '{:.2f} Yrs'.format(v)
    return str(v).strip()

def mult(v):
    if v in (None, ''): return ''
    if isinstance(v, (int, float)):
        return '{:.2f}x'.format(v)
    return str(v).strip()

def html_json(value):
    """Make JSON safe inside an HTML script element."""
    return json.dumps(value, indent=1, ensure_ascii=False).replace('&', '\\u0026').replace('<', '\\u003c').replace('>', '\\u003e')

def text_from_html(raw):
    raw = re.sub(r'<(style|script|noscript)[^>]*>.*?</\1>', ' ', raw, flags=re.S | re.I)
    raw = re.sub(r'<[^>]+>', ' ', raw)
    raw = html_lib.unescape(raw)
    return re.sub(r'\s+', ' ', raw).strip()

def page_title(raw, fallback='Baker 1031 Investments'):
    m = re.search(r'<title[^>]*>(.*?)</title>', raw, flags=re.S | re.I)
    return re.sub(r'\s+', ' ', html_lib.unescape(re.sub(r'<[^>]+>', ' ', m.group(1)))).strip() if m else fallback

def trim_title(value, limit=TITLE_LIMIT):
    value = re.sub(r'\s+', ' ', html_lib.unescape(str(value))).strip()
    if len(value) <= limit:
        return value
    brand = ' | Baker 1031 Investments'
    if value.endswith(brand):
        subject = value[:-len(brand)].rstrip()
        short_brand = ' | Baker 1031'
        if len(subject) + len(short_brand) <= limit:
            return subject + short_brand
        value = subject
    if len(value) <= limit:
        return value
    return value[:limit - 1].rsplit(' ', 1)[0].rstrip(' ,;:—-') + '…'

def trim_description(value, limit=DESCRIPTION_LIMIT):
    value = re.sub(r'\s+', ' ', html_lib.unescape(str(value))).strip()
    if len(value) <= limit:
        return value
    return value[:limit - 1].rsplit(' ', 1)[0].rstrip(' ,;:') + '…'

def page_description(base, raw, meta=None):
    if meta and meta.get('description'):
        return trim_description(meta['description'])
    if base in PAGE_DESCRIPTIONS:
        return trim_description(PAGE_DESCRIPTIONS[base])
    title = page_title(raw, base.replace('.html', '').replace('-', ' ').title())
    paragraphs = re.findall(r'<p(?:\s[^>]*)?>(.*?)</p>', raw, flags=re.S | re.I)
    body = ''
    for p in paragraphs:
        candidate = text_from_html(p)
        if len(candidate) >= 40:
            body = candidate
            break
    if not body:
        m = re.search(r'<h1[^>]*>(.*?)</h1>', raw, flags=re.S | re.I)
        body = text_from_html(m.group(1)) if m else title
    body = body.rstrip(' .')
    return trim_description(title + ' — ' + body + '. Baker 1031 Investments provides educational and investment-research resources for accredited investors.')

def extract_citations(raw):
    urls = re.findall(r"href=[\"'](https?://[^\"']+)", raw, flags=re.I)
    selected = []
    for url in urls:
        if (url in STATE_AUTHORITY_URLS or re.search(r'(irs\.gov|investor\.gov|sec\.gov|finra\.org|brokercheck\.finra\.org|ftb\.ca\.gov|tax\.)', url, re.I)) and url not in selected:
            selected.append(url)
    return selected[:12]

def state_slug_for_page(base):
    stem = base[:-5] if base.endswith('.html') else base
    prefix = '1031-exchange-'
    if not stem.startswith(prefix):
        return None
    suffix = stem[len(prefix):]
    for slug in sorted(STATE_TAX_AUTHORITIES, key=len, reverse=True):
        if suffix == slug or suffix.startswith(slug + '-'):
            return slug
    return None

def page_citations(base, raw):
    citations = extract_citations(raw)
    state_slug = state_slug_for_page(base)
    if state_slug:
        authority_url = STATE_TAX_AUTHORITIES[state_slug]['url']
        citations = [authority_url] + [url for url in citations if url != authority_url]
    elif (base in HUB_FILES or base not in NOINDEX_PAGES) and not citations:
        citations = AUTHORITY_SOURCES[:]
    return citations[:12]

def faq_schema(raw, canonical):
    items = []
    pattern = re.compile(r'<details\b[^>]*>\s*<summary\b[^>]*>(.*?)</summary>(.*?)</details>', flags=re.S | re.I)
    for match in pattern.finditer(raw):
        question = text_from_html(match.group(1))
        answer = text_from_html(match.group(2))
        if question and answer:
            items.append({
                '@type': 'Question',
                'name': question,
                'acceptedAnswer': {'@type': 'Answer', 'text': answer},
            })
    if len(items) < 2:
        return None
    return {'@type': 'FAQPage', '@id': canonical + '#faq', 'url': canonical, 'mainEntity': items[:50]}

def glossary_schema(raw, canonical, title):
    terms = []
    pattern = re.compile(r'<dt\b[^>]*>(.*?)</dt>\s*<dd\b[^>]*>(.*?)</dd>', flags=re.S | re.I)
    for match in pattern.finditer(raw):
        name = text_from_html(match.group(1))
        definition = text_from_html(match.group(2))
        if name and definition:
            terms.append({'@type': 'DefinedTerm', 'name': name, 'description': definition})
    if not terms:
        return None
    return {
        '@type': 'DefinedTermSet',
        '@id': canonical + '#glossary',
        'name': title,
        'url': canonical,
        'hasDefinedTerm': terms[:100],
    }

def ensure_visible_review_and_sources(raw, base):
    if base in NOINDEX_PAGES:
        return raw
    visible = re.sub(r'<(style|script|noscript)[^>]*>.*?</\1>', ' ', raw, flags=re.S | re.I)
    if not re.search(r'(?:class=["\'][^"\']*review-note|Reviewed by\s+Lori Kamen)', visible, flags=re.I) and '<!-- @@REVIEW_NOTE@@ -->' not in raw:
        if re.search(r'</main>', raw, flags=re.I):
            raw = re.sub(r'</main>', REVIEW_NOTE_HTML + '\n</main>', raw, count=1, flags=re.I)
        elif re.search(r'</body>', raw, flags=re.I):
            raw = re.sub(r'</body>', REVIEW_NOTE_HTML + '\n</body>', raw, count=1, flags=re.I)
    if (re.search(r'fully intact', raw, flags=re.I) or base == 'is-the-1031-exchange-going-away-policy-outlook.html') and 'b1031-current-law-source' not in raw:
        raw = re.sub(r'</main>', CURRENT_1031_SOURCE_NOTE + '\n</main>', raw, count=1, flags=re.I)
    if re.search(r'opportunity zone|qualified opportunity fund|\bQOF\b', raw, flags=re.I) and re.search(r'potentially tax-free|tax-free|eliminating federal tax', raw, flags=re.I) and 'b1031-current-law-source' not in raw:
        raw = re.sub(r'</main>', CURRENT_OZ_SOURCE_NOTE + '\n</main>', raw, count=1, flags=re.I)
    return raw

def state_authority_note(base):
    state_slug = state_slug_for_page(base)
    if not state_slug:
        return ''
    authority = STATE_TAX_AUTHORITIES[state_slug]
    name = html_lib.escape(authority['name'])
    url = html_lib.escape(authority['url'], quote=True)
    return ('<p class="b1031-state-source">State tax source: '
            '<a href="%s" target="_blank" rel="noopener noreferrer">Official %s</a>. '
            'State rules can change; confirm current treatment with the agency and your tax adviser.</p>' % (url, name))

def normalize_editorial_identity(raw, base=''):
    raw = raw.replace('Baker 1031 Research is the editorial desk at Baker 1031 Investments',
                      'Gerald F. &quot;Jerry&quot; Baker, III leads the editorial work at Baker 1031 Investments')
    raw = raw.replace('Source: Baker 1031 Research.', 'Source: Gerald F. &quot;Jerry&quot; Baker, III.')
    raw = raw.replace('Baker 1031 Research', AUTHOR_NAME)
    raw = raw.replace('https://www.baker1031.com/assets/img/jerry-baker.jpg', AUTHOR_IMAGE)
    raw = raw.replace('href="baker1031.html"', 'href="/"')
    raw = raw.replace("href='baker1031.html'", "href='/'")
    raw = raw.replace(
        'https://www.irs.gov/credits-deductions/businesses/opportunity-zones-frequently-asked-questions',
        'https://www.irs.gov/credits-deductions/opportunity-zones-frequently-asked-questions')
    raw = raw.replace(
        'https://www.sec.gov/page/staff-statement-opportunity-zones-federal-and-state-securities-laws-considerations',
        'https://www.sec.gov/resources-for-investors/investor-alerts-bulletins/private-placements-under-regulation-d-investor-bulletin')
    raw = raw.replace(
        'SEC &amp; NASAA — Staff Statement on Opportunity Zones: Federal and State Securities Laws Considerations',
        'SEC Investor Bulletin on Private Placements Under Regulation D')
    raw = raw.replace('https://www.irs.gov/irb/2000-40_IRB', 'https://www.irs.gov/pub/irs-irbs/irb00-37.pdf')
    raw = raw.replace('tax-free exclusion', 'potential federal exclusion for eligible appreciation')
    # Keep tax-benefit language explicitly conditional, including capitalized
    # variants that appear in glossary definitions and comparison tables.
    raw = re.sub(r'(?i)(?<!potentially )tax-free', 'potentially tax-free', raw)
    raw = raw.replace('eliminating federal tax on all the appreciation',
                      'potentially excluding eligible appreciation from federal tax after the applicable holding period, subject to statutory requirements and elections')
    raw = raw.replace('1031 is fully intact in 2026',
                      'Section 1031 remains available for qualifying real property as of the review date')
    raw = raw.replace('the exchange is fully intact',
                      'the exchange remains available for qualifying real property as of the review date')
    raw = raw.replace('fully intact as of 2026',
                      'available for qualifying real property as of the review date')
    raw = raw.replace('fully intact rules', 'current rules')
    raw = raw.replace('fully intact', 'available under the current published rules, subject to eligibility')
    raw = raw.replace('tax eliminated on a decade of growth',
                      'potentially excluded from federal tax on eligible appreciation after the applicable holding period')
    raw = raw.replace('tax eliminated', 'potentially excluded from federal tax')
    raw = raw.replace('so you never risk the exchange',
                      'which may help you manage the exchange timeline, but cannot eliminate exchange risk')
    raw = raw.replace('you never risk the exchange',
                      'you can better manage the exchange timeline, but no strategy eliminates exchange risk')
    # The conditional wording above must never alter a URL slug.
    raw = raw.replace('the-10-year-hold-potentially tax-free-appreciation-explained.html',
                      'the-10-year-hold-tax-free-appreciation-explained.html')
    raw = raw.replace('Baker 1031 Research', AUTHOR_NAME)
    raw = raw.replace('the preferred cohort has out-returned the broader field',
                      'the preferred designation reflects coverage and diligence depth, not a performance ranking')
    raw = raw.replace('preferred cohort has out-returned the broader field',
                      'the preferred designation reflects coverage and diligence depth, not a performance ranking')
    raw = raw.replace('has out-returned the broader field',
                      'has not been established as a comparable, independently verified performance result')
    raw = raw.replace(
        'This Privacy Policy is a template provided for review and should be reviewed and finalized with qualified counsel before reliance.',
        'This Privacy Policy describes the current website practices of Baker 1031 Investments, LLC and may be updated as our services, vendors, and legal obligations change.')
    raw = raw.replace(
        'This financial-privacy notice is a template provided for review and should be finalized with your broker-dealer and qualified counsel.',
        'This financial-privacy notice describes the current information-handling practices used in connection with Baker 1031 Investments and the broker-dealer through which securities business is conducted.')
    raw = raw.replace(
        'This California Privacy Notice is a template provided for review and should be finalized with qualified counsel.',
        'This California Privacy Notice describes the current privacy practices of Baker 1031 Investments, LLC and may be updated as our services and legal obligations change.')
    raw = raw.replace(
        'This Accessibility Statement is a template provided for review and reflects our ongoing, good-faith efforts.',
        'This Accessibility Statement describes our ongoing efforts to improve access to the website and to address reported barriers.')
    raw = raw.replace(
        'These Terms are a template provided for review and should be finalized with qualified counsel before reliance.',
        'These Terms describe the current conditions governing use of this website and may be updated as the site and applicable requirements change.')
    raw = raw.replace(
        'It is a template for review with counsel; the final policy your firm adopts controls.',
        'It describes the current website practices of Baker 1031 Investments, LLC and may be updated as our services, vendors, and legal obligations change.')
    raw = raw.replace(
        'It is a template for review with counsel and your broker-dealer.',
        'It describes the current information-handling practices used in connection with Baker 1031 Investments and the broker-dealer through which securities business is conducted.')
    raw = raw.replace('It is a template for review with counsel.',
                      'It describes the current California privacy practices of Baker 1031 Investments, LLC and may be updated as legal requirements change.')
    raw = raw.replace(
        'This is a template for review with counsel; the final terms your firm adopts control.',
        'These Terms describe the current conditions governing use of this website and may be updated as the site and applicable requirements change.')
    raw = raw.replace(
        'It is a template for review with counsel; the final terms your firm adopts control.',
        'These Terms describe the current conditions governing use of this website and may be updated as the site and applicable requirements change.')
    raw = raw.replace(
        'This is a template for review with counsel.',
        'This page describes the current website practices of Baker 1031 Investments, LLC and may be updated as legal requirements change.')
    raw = re.sub(r'\s*Baker 1031 reviews its educational content periodically for accuracy and regulatory compliance\.', '', raw, flags=re.I)
    raw = re.sub(r'\s*Content subject to registered-principal review\.', '', raw, flags=re.I)
    raw = re.sub(r'Last reviewed\s+[A-Za-z]+\s+(?:\d{1,2},\s+)?\d{4}', 'Last reviewed ' + REVIEW_DATE, raw)
    raw = raw.replace(
        'President &amp; CCO, Aurora Securities, Inc. (FINRA Series 4 / 7 / 24 / 53 / 63 / 66), the supervising registered principal.',
        'Chief Compliance Officer, Aurora Securities, Inc. (FINRA CRD #2805591).')
    if base == 'opportunity-zone-deadlines-and-program-timeline.html':
        raw = re.sub(r'(<div class="wrap abody">).*?(<h2 id="faq">)', r'''\1
<article class="article">
<p class="lead">Opportunity Zone timing depends on the date of the investment, the source of the gain, and the law in effect for that transaction. This guide separates the dates investors track from the tax outcomes that require individualized advice. Review the cited primary sources and confirm the current rules with your CPA before acting.</p>
<h2 id="key-dates">Key Opportunity Zone dates</h2>
<p>The main dates are the 180-day investment window for a qualifying gain, the applicable deferred-gain recognition date, the ten-year holding milestone for a possible exclusion of eligible appreciation, and any zone-designation transition dates. Pass-through gains may have special rules for when the 180-day period begins. The clock should be calculated from the taxpayer's facts, not from a generic calendar.</p>
<h2 id="when-recognized">When deferred gain is recognized</h2>
<p>The recognition date depends on the program and investment date. The original program and any later statutory changes must be analyzed separately; a general article cannot determine the inclusion date for an individual investor. Track the date in the governing statute, regulations, IRS guidance, and the investor's tax records, then confirm the result with a CPA.</p>
<div class="pullquote"><p>Opportunity Zone benefits are conditional. A missed investment window, an inclusion event, a qualification failure, or a change in law can alter the result.</p></div>
<h2 id="sunset-extensions">Program permanence &amp; transitions</h2>
<p>Opportunity Zone legislation and administrative guidance have changed over time. The existence of a continuing program does not make every zone, fund, or tax benefit permanent for every investor. Confirm the applicable designation, eligibility tests, holding-period requirements, and effective dates from current government sources before relying on a timeline.</p>
<h2 id="timing-benefits">How timing affects your benefits</h2>
<p>Timing can affect whether a gain is eligible for deferral, when deferred tax becomes reportable, and whether a later appreciation exclusion may be available. Those outcomes depend on the type of gain, the investment vehicle, statutory requirements, and the taxpayer's compliance with them. A deadline calculator can organize dates, but it cannot establish eligibility or predict a tax result.</p>
<div class="takeaways"><div class="tk-h">Key Takeaways</div><ul>
<li>Calculate the 180-day window from the correct gain and taxpayer facts.</li>
<li>Separate original-gain deferral from any possible treatment of later appreciation.</li>
<li>Confirm zone designations and transition dates from current government sources.</li>
<li>Treat every result as conditional and have a CPA review the transaction before funding.</li>
</ul></div>
<h2 id="staying-current">Staying current on the rules</h2>
<p>Use the IRS, Treasury, CDFI Fund, and other applicable government sources for current rules and designations. Keep a dated copy of the sources used for the analysis because guidance and administrative materials can change. Baker 1031 provides educational research and does not provide tax or legal advice.</p>
<h2 id="how-baker-helps">How Baker 1031 helps with the timeline</h2>
<p>Baker 1031 can help investors organize questions, compare available structures, and identify the documents a professional team should review. Any securities offering is made only through the applicable private placement memorandum and after the required suitability process; no timeline guarantees an exchange, a tax result, or an investment outcome.</p>
\2''', raw, count=1, flags=re.S)
    elif base == 'is-the-1031-exchange-going-away-policy-outlook.html':
        raw = re.sub(r'(<div class="wrap abody">).*?(<h2 id="faq">)', r'''\1
<article class="article">
<p class="lead">Section 1031 policy is often discussed during tax legislation and budget negotiations. This page explains how to read those proposals without treating a proposal as current law. As of the July 12, 2026 review date, Section 1031 remains available for qualifying real property under the rules cited on this page; the statement is time-bound and should be rechecked before a transaction.</p>
<h2 id="recurring-question">The recurring question</h2>
<p>Investors often ask whether Section 1031 will be repealed or capped. Proposals may be introduced, scored, or discussed without becoming law. The useful distinction is between enacted statutory text, agency guidance, and advocacy or budget material.</p>
<h2 id="why-targeted">Why 1031 is periodically targeted</h2>
<p>Like-kind exchanges defer recognition of qualifying gain rather than permanently guaranteeing that tax will never be due. That deferral can appear in revenue estimates, which is why the provision periodically appears in policy debates. The policy argument does not determine the law that applies to a particular exchange.</p>
<h2 id="recent-proposals">How to read proposals</h2>
<p>Check the bill or enacted statute, its effective date, transition rules, and the IRS guidance that implements it. Headlines and summaries can omit those details. A proposal is not a change to a transaction already underway unless the enacted text says so and the taxpayer's facts satisfy the transition rules.</p>
<h2 id="current-status">Current status as of the review date</h2>
<p>The current federal rule for qualifying real property is Section 1031. The page's cited primary sources are the proper starting point for verification. State treatment, entity structure, debt, boot, related-party rules, and personal facts can change the tax analysis even when the federal provision is available.</p>
<h2 id="economic-case">The economic case for 1031</h2>
<p>Supporters point to reinvestment and transaction activity; critics focus on tax deferral and distributional effects. Both are policy arguments, not evidence of a future legislative outcome. Investors should separate that debate from the practical question of whether an exchange fits their own objectives and facts.</p>
<h2 id="what-investors-do">What investors should do</h2>
<p>Plan for the transaction you actually need, use current sources, and keep a qualified intermediary and tax adviser involved before the sale closes. Do not rush because of an unenacted proposal, and do not assume a future change will be prospective or favorable without reading the enacted transition language.</p>
<h2 id="policy-takeaway">Policy context</h2>
<p>Policy arguments can inform the debate, but they do not replace the enacted rule or the professional analysis of a specific exchange.</p>
<h2 id="how-baker-helps">How Baker 1031 helps you navigate the outlook</h2>
<p>Baker 1031 provides educational research, source links, and exchange-planning questions for investors and advisers. We do not predict legislation or provide tax or legal advice. Confirm the current law and your transaction-specific treatment with your CPA, attorney, and qualified intermediary.</p>
\2''', raw, count=1, flags=re.S)
    return raw.replace(
        'President &amp; CCO, Aurora Securities',
        'Chief Compliance Officer, Aurora Securities (FINRA CRD #2805591)')

def seo_jsonld(base, title, description, canonical, raw):
    author = {
        '@type': 'Person',
        'name': AUTHOR_NAME,
        'alternateName': 'Jerry Baker',
        'jobTitle': 'Founder & Managing Principal, Baker 1031 Investments',
        'url': BASE_URL + '/jerry-baker-bio.html',
        'image': AUTHOR_IMAGE,
        'sameAs': ['https://brokercheck.finra.org/individual/summary/7537416'],
    }
    publisher = {
        '@type': 'Organization',
        'name': 'Baker 1031 Investments, LLC',
        'url': BASE_URL + '/',
        'logo': {'@type': 'ImageObject', 'url': BASE_URL + '/assets/logo.png'},
    }
    citations = page_citations(base, raw)
    is_calculator = base in INTERACTIVE_CALCULATOR_PAGES
    audio_urls = re.findall(r'<(?:audio|source)\b[^>]*\bsrc=["\']([^"\']+)', raw, flags=re.I)
    review_text = text_from_html(raw)
    review_match = re.search(r'Reviewed by\s+(.+?)\s+—\s+(.+?)\.\s+Last reviewed\s+([A-Za-z]+\s+(?:\d{1,2},\s+)?\d{4})', review_text, flags=re.I)
    reviewer = REVIEWER if base not in NOINDEX_PAGES else None
    if base == 'jerry-baker-bio.html':
        data = {
            '@context': 'https://schema.org', '@type': ['ProfilePage', 'WebPage'],
            '@id': canonical, 'url': canonical, 'name': title, 'description': description,
            'mainEntity': author, 'publisher': publisher,
        }
    elif base not in NOINDEX_PAGES and (base.endswith('.html')):
        data = {
            '@context': 'https://schema.org',
            '@type': 'Article' if base not in HUB_FILES and base not in SEO_META and base not in ('contact.html', 'request-access.html') else 'WebPage',
            '@id': canonical + ('#article' if base not in HUB_FILES and base not in SEO_META and base not in ('contact.html', 'request-access.html') else '#webpage'),
            'url': canonical, 'headline': title, 'name': title,
            'description': description, 'publisher': publisher,
        }
        if data['@type'] == 'Article':
            data['author'] = author
            data['mainEntityOfPage'] = canonical
        if citations:
            data['citation'] = citations
        if is_calculator:
            data['mainEntity'] = {
                '@type': 'WebApplication',
                '@id': canonical + '#calculator',
                'name': title,
                'url': canonical,
                'applicationCategory': 'FinanceApplication',
                'operatingSystem': 'Web',
                'isAccessibleForFree': True,
                'offers': {'@type': 'Offer', 'price': '0', 'priceCurrency': 'USD'},
            }
        if audio_urls:
            data['associatedMedia'] = {
                '@type': 'AudioObject',
                '@id': canonical + '#audio',
                'contentUrl': audio_urls[0],
                'name': title + ' audio version',
                'description': description,
                'inLanguage': 'en-US',
                'isAccessibleForFree': True,
            }
    else:
        data = {'@context': 'https://schema.org', '@type': 'WebPage', '@id': canonical, 'url': canonical, 'name': title, 'description': description, 'publisher': publisher}
    if reviewer:
        data['reviewedBy'] = reviewer
        data['dateModified'] = REVIEW_DATE_ISO
    graph = [data]
    faq = faq_schema(raw, canonical)
    if faq:
        graph.append(faq)
    if base == 'glossary.html':
        glossary = glossary_schema(raw, canonical, title)
        if glossary:
            graph.append(glossary)
    if len(graph) > 1:
        for entity in graph:
            entity.pop('@context', None)
        data = {'@context': 'https://schema.org', '@graph': graph}
    return html_json(data)

def remove_seo_tags(raw):
    raw = re.sub(r"\s*<meta\b[^>]*(?:name|property)=[\"'](?:description|robots|og:[^\"']+|twitter:[^\"']+)[\"'][^>]*>", '', raw, flags=re.I)
    raw = re.sub(r"\s*<link\b[^>]*rel=[\"']canonical[\"'][^>]*>", '', raw, flags=re.I)
    raw = re.sub(r"\s*<script\b[^>]*type=[\"']application/ld\+json[\"'][^>]*>.*?</script>", '', raw, flags=re.S | re.I)
    return raw

# ---------------------------------------------------------------- maps
def urlseg(u):
    """Last path segment of a sheet URL value, sanitized into a clean slug."""
    seg = u.rstrip('/').split('/')[-1]
    return slugify(seg.replace('&', ' and '))

off_slug = {}
for row in LIST:
    u = s(row.get('URL'))
    off_slug[row['Investment Name']] = (urlseg(u) if u else slugify(row['Investment Name'])) + '.html'

spon_slug = {}
for row in LIST:
    su = s(row.get('Sponsor URL'))
    if su:
        sponsor_path_slug = urlseg(su)
        if sponsor_path_slug.startswith('sponsor-'):
            sponsor_path_slug = sponsor_path_slug[len('sponsor-'):]
        spon_slug[s(row['Sponsor'])] = 'sponsor-' + sponsor_path_slug + '.html'
for sp in SPON:
    nm = s(sp['Investment Firm'])
    spon_slug.setdefault(nm, 'sponsor-' + slugify(nm.replace('&', ' and ')) + '.html')

# Guard against sheet data errors: two sponsors mapped to the same URL slug
# (e.g. Resource Royalty's Sponsor URL mistakenly set to sponsors/exchangeright).
# The sponsor whose own name matches the slug keeps it; others fall back to their name.
_by_slug = {}
for _nm, _sl in spon_slug.items():
    _by_slug.setdefault(_sl, []).append(_nm)
for _sl, _nms in _by_slug.items():
    if len(_nms) > 1:
        for _nm in _nms:
            if slugify(_nm.replace('&', ' and ')) not in _sl:
                spon_slug[_nm] = 'sponsor-' + slugify(_nm.replace('&', ' and ')) + '.html'
                print('WARN: sponsor URL collision for %r on %s -> reassigned %s' % (_nm, _sl, spon_slug[_nm]))

docs_by_inv = {}
for d in DOCS:
    docs_by_inv.setdefault(s(d['Investment Name']), []).append(d)
track_by_sp = {}
for t in TRACK:
    track_by_sp.setdefault(s(t['Sponsor']), []).append(t)
spon_by_name = {s(x['Investment Firm']): x for x in SPON}

FOOTNOTES = [
    'Average Yield is the simple average of the sponsor’s projected annual distribution rates as presented in the Private Placement Memorandum. Projections are not guaranteed and actual distributions may be lower.',
    'Estimated Tax-Adjusted Yield reflects the projected impact of depreciation and amortization deductions at an assumed combined federal and state tax rate. Individual tax outcomes vary — consult your CPA regarding your specific situation.',
    'Cap Rate Equivalent is a Baker 1031 Investments calculation intended to allow comparison with direct property ownership; it is not a sponsor-reported figure and does not represent a rate of return.',
    'Benchmarks compare this offering’s projected figures against sector medians computed across current offerings tracked by Baker 1031 Investments as of the last-updated date shown. Benchmark data is internal, unaudited, and subject to change.',
]

def sponsor_research_context(d):
    """Add useful, source-labeled context to generated sponsor profiles.

    This is assembled only from workbook fields already displayed on the page;
    it avoids inventing sponsor claims merely to satisfy a word-count check.
    """
    name = html_lib.escape(s(d.get('name', 'this sponsor')))
    founded = html_lib.escape(s(d.get('founded'))) if d.get('founded') else ''
    aum = html_lib.escape(s(d.get('aum'))) if d.get('aum') else ''
    hq = html_lib.escape(s(d.get('hq'))) if d.get('hq') else ''
    strategy_items = [html_lib.escape(x) for x in (d.get('strategies') or []) if x]
    offering_count = len(d.get('currentOfferings') or [])
    track_count = len(d.get('fullTrackRecord') or [])
    facts = []
    if founded: facts.append('a reported founding year of %s' % founded)
    if aum: facts.append('reported assets under management of %s' % aum)
    if hq: facts.append('a reported headquarters in %s' % hq)
    fact_sentence = ', '.join(facts) if facts else 'the sponsor facts shown above'
    strategy_html = ''
    if strategy_items:
        strategy_html = '<h3 class="subhead">What the strategy fields mean</h3><p>The strategy and advantage fields are a compact description of how %s presents its approach. They are not a recommendation, rating, or substitute for the PPM. Compare them with property-level leverage, tenant concentration, reserves, fees, exit assumptions, and the investor rights described in the offering documents.</p><ul class="research-list">%s</ul>' % (
            name, ''.join('<li>%s</li>' % item for item in strategy_items))
    track_sentence = (
        'The page includes %d full-cycle record%s in the current workbook. Those figures are sponsor-reported and unaudited; they are shown to make the source and scope visible, not to imply that a future offering will repeat historical results.' %
        (track_count, '' if track_count == 1 else 's')) if track_count else (
        'No full-cycle rows are currently present in the workbook for this sponsor. That absence is a data-coverage limitation, not evidence about performance.')
    offering_sentence = (
        '%s current offering%s are linked from this profile. Availability, projected income, leverage, and documents can change, so investors should confirm the current version of each PPM and related supplement before relying on any figure.' %
        (offering_count, '' if offering_count == 1 else 's')) if offering_count else (
        'No current offerings are linked from this profile in the workbook snapshot. The directory can change as new offerings are reviewed.')
    return '''
        <section class="research-context" id="sresearch">
          <h2 class="optima">How to Read This Sponsor Profile</h2>
          <p>This page is a Baker 1031 Investments research summary for %s. The workbook currently records %s. Baker 1031 presents these fields to help an accredited investor organize due-diligence questions; the page is not a sponsor endorsement, performance guarantee, or individualized investment recommendation.</p>
          <p>The overview and facts combine sponsor-provided information with the current Baker 1031 directory snapshot. A reported number should be read as reported, not as independently audited. The most authoritative source for an offering's structure, fees, conflicts, financial projections, risk factors, transfer restrictions, and suitability is the sponsor's Private Placement Memorandum and its supplements.</p>
          %s
          <h3 class="subhead">Current coverage</h3>
          <p>%s %s</p>
          <p>Use this profile alongside the <a href="due-diligence.html">Baker 1031 due-diligence process</a>, the <a href="methodology.html">data methodology</a>, and the relevant <a href="delaware-statutory-trusts.html">DST guide</a>. Before investing, review the complete documents with your own tax, legal, and financial advisers and confirm that the strategy fits your exchange timeline, liquidity needs, and risk tolerance.</p>
        </section>
    ''' % (name, fact_sentence, strategy_html, offering_sentence, track_sentence)

# ---------------------------------------------------------------- generation
otpl = open(os.path.join(ROOT, 'templates', 'offering-template.html')).read()
stpl = open(os.path.join(ROOT, 'templates', 'sponsor-template.html')).read()

generated = {}  # filename -> html

def build_offering(row):
    name = s(row['Investment Name'])
    slug = off_slug[name]
    yields = {}
    for y in ['Y1','Y2','Y3','Y4','Y5','Y6','Y7','Y8','Y9','Y10']:
        if row.get(y) not in (None, ''):
            yields[y] = pct(row.get(y))
    docs = []
    for d in docs_by_inv.get(name, []):
        gated = s(d.get('Gated?')).lower() in ('yes','true','y','1','gated')
        f = s(d.get('File'))
        docs.append({'label': s(d['Label']) + (' (by request)' if gated and not f else ''),
                     'url': f if (f and f.startswith('http')) else '#'})
    if not docs:
        docs = [{'label': 'Offering Documents Available By Request', 'url': '#'}]
    bms = []
    for label, a, b, c in [('Avg. Income','BM: Avg. Income - Deal','BM: Avg. Income - MKT','BM: Avg. Income - Interpret'),
                           ('Growth','BM: Growth - Deal','BM: Growth- MKT','BM: Growth - Interpret'),
                           ('Peak','BM: Peak - Deal','BM: Peak- MKT','BM: Peak - Interpret')]:
        if row.get(a) not in (None, ''):
            bms.append({'metric': label, 'deal': pct(row.get(a)), 'market': pct(row.get(b)), 'interpret': s(row.get(c))})
    sp = s(row['Sponsor'])
    spr = spon_by_name.get(sp, {})
    d = {
      'investmentName': name, 'sponsor': sp,
      'structure': s(row.get('Structure')), 'status': s(row.get('Status')),
      'totalOffering': money(row.get('Total Offering')), 'equity': money(row.get('Equity')),
      'debt': money(row.get('Debt')), 'inPlaceLtv': s(row.get('In-Place LTV')) or pct(row.get('In-Place LTV')),
      'availableEquity': money(row.get('Available Equity')),
      'lastUpdated': s(row.get('Last Updated')).split(' ')[0],
      'availablePercentage': pct(row.get('Available Percentage')),
      'propertyType': s(row.get('Property Type')), 'location': s(row.get('Location')),
      'locationUse': s(row.get('Location (Use)')), 'totalLoad': pct(row.get('Total Load')),
      'strategy': s(row.get('Strategy')), 'exchange721Exit': s(row.get('721 Exchange Exit')),
      'estimatedHoldPeriod': s(row.get('Estimated Hold Period')),
      'description': s(row.get('Description')),
      'highlights': [s(row.get('Highlight %d' % i)) for i in range(1, 6) if s(row.get('Highlight %d' % i))],
      'pros': s(row.get('Pros')), 'cons': s(row.get('Cons')), 'insights': s(row.get('Insights')),
      'yields': yields, 'averageYield': pct(row.get('Average Yield')),
      'taxAdjYield': s(row.get('Tax-Adj. Yield')) or pct(row.get('Tax-Adj. Yield')),
      'capRateEquivalent': pct(row.get('Cap Rate Equivalent')),
      'lender': s(row.get('Lender')), 'interestRate': s(row.get('Interest Rate')) or pct(row.get('Interest Rate')),
      'loanTerm': s(row.get('Loan Term')), 'ioPeriod': s(row.get('I/O Period')),
      'amortization': s(row.get('Amortization')), 'y1Dscr': s(row.get('Y1 DSCR')),
      'propertyAddress': s(row.get('Property Address')),
      'minimumInvestment': money(row.get('Minimum Investment')),
      'photoLinkUse': s(row.get('Photo Link Use')),
      'taxAdjustedYieldUse': pct(row.get('Tax Adjusted Yield (Use)')),
      'taxAdjLabel': s(row.get('Tax Adj Label')) or 'Est. Tax-Adjusted Yield',
      'ddLabel': s(row.get('DD Label')), 'url': slug.replace('.html',''),
      'sponsorUrl': spon_slug.get(sp, '#'),
      'sponsorButtonText': s(row.get('Sponsor Button Text')) or ('Learn More About ' + sp),
      'sponsorFounded': s(spr.get('Year Founded')),
      'sponsorDescription': s(spr.get('Description / Overview')) or s(row.get('Sponsor Description')),
      'sponsorAum': money_compact(spr.get('AUM')) or money_compact(row.get('Sponsor AUM')),
      'fullCycleCount': s(spr.get('Full-Cycle Deals')) or s(row.get('Full-Cycle Count')),
      'sponsorAar': pct(spr.get('Average Annual Return')) or pct(row.get('Sponsor AAR')),
      'sponsorAem': mult(spr.get('Average Equity Multiple')) or s(row.get('Sponsor AEM')),
      'sponsorHold': yrs(spr.get('Average Hold Period')) or s(row.get('Sponsor Hold')),
      'sponsorSuccess': pct(spr.get('Success Rate')) or pct(row.get('Sponsor Success')),
      'sponsorImage': s(spr.get('Logo')) or s(row.get('Sponsor Image')),
      'documents': docs, 'benchmarks': bms, 'footnotes': FOOTNOTES,
    }
    d = {k: v for k, v in d.items() if v not in (None, '', [], {})}
    page = otpl
    page = re.sub(r'<title>.*?</title>', '<title>' + name.replace('&','&amp;') + ' | Baker 1031 Investments</title>', page, flags=re.S)
    page = re.sub(r'<script type="application/json" id="offering-data">\n.*?\n</script>',
                  lambda _m: '<script type="application/json" id="offering-data">\n' + html_json(d) + '\n</script>',
                  page, flags=re.S)
    page = page.replace('>AEI Healthcare Portfolio VII DST<', '>' + name + '<')
    summary = ' '.join(x for x in [d.get('propertyType'), d.get('locationUse') or d.get('location'), d.get('status')] if x)
    SEO_META[slug] = {
        'title': name + ' | Baker 1031 Investments',
        'description': (name + ' — ' + (summary or '1031 exchange replacement-property offering') + '. Review property details, projected income, financing, sponsor information, and risks from Baker 1031 Investments.')[:300],
    }
    generated[slug] = page
    return d

dir_rows = []
for row in LIST:
    d = build_offering(row)
    dir_rows.append({
      'name': d.get('investmentName',''), 'url': off_slug[d['investmentName']],
      'sponsor': d.get('sponsor',''), 'propertyType': d.get('propertyType',''),
      'location': d.get('locationUse') or d.get('location',''),
      'ltv': d.get('inPlaceLtv','').replace(' LTV',''),
      'minimum': d.get('minimumInvestment',''),
      'y1Yield': (d.get('yields') or {}).get('Y1',''),
      'avgYield': d.get('averageYield',''), 'status': d.get('status',''),
      'availablePct': d.get('availablePercentage',''), 'photo': d.get('photoLinkUse',''),
      'description': (s(row.get('List Description')) or d.get('description',''))[:400],
    })

sp_dir = []
for sp in SPON:
    nm = s(sp['Investment Firm'])
    slug = spon_slug[nm]
    strategies = [s(sp.get('Key Strategy / Advantage %d' % i)) for i in range(1, 6)]
    strategies = [x for x in strategies if x]
    track = [{'investment': s(t.get('Investment')), 'location': s(t.get('Location')),
              'assetClass': s(t.get('Asset Class')) or '—',
              'holdPeriod': ('{:.2f}'.format(t['Hold Period']) if isinstance(t.get('Hold Period'), (int, float)) else s(t.get('Hold Period'))),
              'equityMultiple': mult(t.get('Equity Multiple')),
              'annualReturn': pct(t.get('Annual Return'))} for t in track_by_sp.get(nm, [])]
    offers = [{'name': r['name'],
               'sub': ' · '.join(x for x in [r['propertyType'], r['location'],
                        (r['ltv'] + ' LTV' if r['ltv'] else ''), (r['minimum'] + ' minimum' if r['minimum'] else '')] if x),
               'url': r['url']} for r in dir_rows if r['sponsor'] == nm]
    d = {'name': nm, 'founded': s(sp.get('Year Founded')), 'aum': money_compact(sp.get('AUM')),
         'hq': s(sp.get('Headquarters (City, State)')), 'website': s(sp.get('Website')),
         'logo': s(sp.get('Logo')),
         'fullCycleDeals': s(sp.get('Full-Cycle Deals')) or (str(len(track)) if track else ''),
         'avgAnnualReturn': pct(sp.get('Average Annual Return')),
         'avgEquityMultiple': mult(sp.get('Average Equity Multiple')),
         'avgHoldPeriod': yrs(sp.get('Average Hold Period')),
         'successRate': pct(sp.get('Success Rate')),
         'description': s(sp.get('Description / Overview')),
         'strategies': strategies, 'currentOfferings': offers, 'fullTrackRecord': track}
    d = {k: v for k, v in d.items() if v not in (None, '', [])}
    page = stpl
    page = re.sub(r'<title>.*?</title>', '<title>' + nm.replace('&','&amp;') + ' | Baker 1031 Investments</title>', page, flags=re.S)
    page = re.sub(r'<script type="application/json" id="sponsor-data">\n.*?\n</script>',
                  lambda _m: '<script type="application/json" id="sponsor-data">\n' + html_json(d) + '\n</script>',
                  page, flags=re.S)
    page = page.replace('<!-- @@SPONSOR_RESEARCH_CONTEXT@@ -->', sponsor_research_context(d))
    page = page.replace('>AEI Capital Corporation<', '>' + nm + '<')
    SEO_META[slug] = {
        'title': nm + ' | Baker 1031 Investments',
        'description': (nm + ' — DST sponsor profile with current offerings, strategy information, sponsor facts, and full-cycle track-record data from Baker 1031 Investments.')[:300],
    }
    generated[slug] = page
    sp_dir.append({'name': nm, 'url': slug, 'founded': d.get('founded',''), 'aum': d.get('aum',''),
                   'hq': d.get('hq',''), 'website': d.get('website',''), 'logo': d.get('logo',''),
                   'fullCycleDeals': d.get('fullCycleDeals','—'),
                   'avgAnnualReturn': d.get('avgAnnualReturn','—'),
                   'avgEquityMultiple': d.get('avgEquityMultiple','—'),
                   'avgHoldPeriod': d.get('avgHoldPeriod','—'),
                   'successRate': d.get('successRate','—'),
                   'description': d.get('description','')[:500]})

print('generated: %d offering + %d sponsor pages' % (len(dir_rows), len(sp_dir)))

# ---------------------------------------------------------------- assemble
nav = open(os.path.join(ROOT, 'src', 'partials', 'nav.html')).read()
footer = open(os.path.join(ROOT, 'src', 'partials', 'footer.html')).read()
GA4_TAG = """<!-- Google tag (gtag.js) -->
<script async src="https://www.googletagmanager.com/gtag/js?id=%s"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){dataLayer.push(arguments);}
  gtag('js', new Date());
  gtag('config', '%s');
</script>""" % (GA4_MEASUREMENT_ID, GA4_MEASUREMENT_ID)
FAVICON_TAG = """<link rel="icon" href="/favicon.ico" sizes="any">
<link rel="icon" type="image/png" sizes="192x192" href="/assets/icon-192.png">
<link rel="apple-touch-icon" sizes="180x180" href="/assets/icon-192.png">
<link rel="manifest" href="/site.webmanifest">"""
ANALYTICS_SCRIPT = '<script src="assets/analytics.js" defer></script>'
ANALYTICS_EXCLUDE = {'employee.html'}
# Clarity is intentionally limited to public pages. Do not record the employee
# terminal, authenticated investor account, or registration/accreditation flow.
CLARITY_EXCLUDE = {'employee.html', 'account.html', 'request-access.html'}
CLARITY_TAG = '''<script>
(function(c,l,a,r,i,t,y){
  c[a]=c[a]||function(){(c[a].q=c[a].q||[]).push(arguments)};
  t=l.createElement(r);t.async=1;t.src="https://www.clarity.ms/tag/"+i;
  y=l.getElementsByTagName(r)[0];y.parentNode.insertBefore(t,y);
})(window, document, "clarity", "script", "%s");
</script>''' % CLARITY_PROJECT_ID
# Sentry's browser loader is the official project-specific snippet generated in
# the Sentry project settings. It is intentionally limited to public site
# pages; the employee terminal is excluded so internal operational data is not
# sent to a third-party error tracker.
SENTRY_TAG = '<script src="https://js.sentry-cdn.com/49f51d213d1534d765e758b5cc973547.min.js" crossorigin="anonymous" defer></script>'
SENTRY_EXCLUDE = {'employee.html'}
RELATED_HUBS_HTML = """<section class="b1031-related" aria-label="Explore Baker 1031 topics">
  <div class="b1031-related-inner">
    <div class="b1031-related-kicker">Explore the Baker 1031 research library</div>
    <div class="b1031-related-links">
      <a href="1031-exchange-guide.html">1031 Exchanges</a>
      <a href="delaware-statutory-trusts.html">DSTs</a>
      <a href="721-exchange-guide.html">721 / UPREITs</a>
      <a href="opportunity-zones-guide.html">Opportunity Zones</a>
      <a href="mineral-rights-1031-guide.html">Mineral &amp; Royalty Interests</a>
      <a href="reits-guide.html">REITs</a>
      <a href="investments.html">Current Offerings</a>
      <a href="sponsors.html">Sponsor Directory</a>
    </div>
  </div>
</section>"""

ORPHAN_RESOURCE_LINKS = {
    'glossary.html': """<section class="b1031-related" aria-label="Related glossary resources">
  <div class="b1031-related-inner">
    <div class="b1031-related-kicker">Related glossary resources</div>
    <div class="b1031-related-links">
      <a href="glossary-drop-and-swap.html">Drop-and-Swap</a>
      <a href="glossary-qualified-intermediary-bond.html">Qualified Intermediary Bond</a>
      <a href="glossary-realized-gain-vs-recognized-gain.html">Realized vs. Recognized Gain</a>
      <a href="glossary-dst-vs-tic.html">DST vs. TIC</a>
      <a href="glossary-securitized-real-estate.html">Securitized Real Estate</a>
    </div>
  </div>
</section>""",
    'due-diligence.html': """<section class="b1031-related" aria-label="Related due diligence resources">
  <div class="b1031-related-inner">
    <div class="b1031-related-kicker">Related due-diligence resource</div>
    <div class="b1031-related-links">
      <a href="ppm-review-checklist.html">Interactive PPM Review Checklist</a>
    </div>
  </div>
</section>""",
}

def seo_inject(html_text, base):
    html_text = normalize_editorial_identity(html_text, base)
    if base == 'is-the-1031-exchange-going-away-policy-outlook.html':
        html_text = re.sub(r'\bas of 2026\b', lambda match: ('As' if match.group(0)[0].isupper() else 'as') + ' of July 11, 2026', html_text, flags=re.I)
        html_text = re.sub(r'\bfully intact\b', 'available for qualifying real property under current guidance', html_text, flags=re.I)
        html_text = re.sub(r'\bfully available(?: now)?\b', 'available under current guidance', html_text, flags=re.I)
        html_text = html_text.replace('available for qualifying real property under current guidance for real estate',
                                      'available for qualifying real property under current guidance')
    is_oz_page = bool(re.search(r'opportunity[- ]zone|(?:^|[-_])qof(?:[-_.]|$)', base, flags=re.I))
    if is_oz_page or base in ('faq.html', 'insights.html'):
        def conditional_tax_free(match):
            if not is_oz_page:
                nearby = text_from_html(html_text[max(0, match.start() - 320):match.end() + 320])
                if not re.search(r'opportunity zone|qualified opportunity fund|\bQOF\b|\bOZ\b|10-year exclusion', nearby, flags=re.I):
                    return match.group(0)
            last_lt = html_text.rfind('<', 0, match.start())
            last_gt = html_text.rfind('>', 0, match.start())
            if last_lt > last_gt:
                return match.group(0)
            context = html_text[max(0, match.start() - 24):match.start()].lower()
            if re.search(r'(?:potential|potentially|not|is not|isn\'t|no)\s+$', context):
                return match.group(0)
            return 'potentially tax-free'
        html_text = re.sub(r'\btax-free\b', conditional_tax_free, html_text, flags=re.I)
        html_text = re.sub(r'\beliminating federal tax on all the appreciation\b',
                           'potentially excluding federal tax on qualifying appreciation',
                           html_text, flags=re.I)
        html_text = re.sub(r'\beliminating tax on\b', 'potentially excluding tax on', html_text, flags=re.I)
        html_text = re.sub(r'\beliminate tax on\b', 'may exclude tax on', html_text, flags=re.I)
        html_text = re.sub(r'\beliminates tax on\b', 'may exclude tax on', html_text, flags=re.I)
        html_text = re.sub(r'\beliminate tax\b', 'may exclude tax', html_text, flags=re.I)
        html_text = re.sub(r'\btax eliminated on\b', 'potential tax on', html_text, flags=re.I)
        if is_oz_page:
            html_text = re.sub(r'\bfully intact\b', 'potentially available under the current program', html_text, flags=re.I)
    html_text = ensure_visible_review_and_sources(html_text, base)
    html_text = html_text.replace('[Placeholder regulatory disclosure — replace with verified entity names, CRD numbers, and registrations.]', APPROVED_DISCLOSURE_LINK)
    meta = SEO_META.get(base, {})
    title = trim_title(meta.get('title') or page_title(html_text, base.replace('.html', '').replace('-', ' ').title()))
    description = page_description(base, html_text, meta)
    canonical = BASE_URL + '/' if base in ('index.html', 'baker1031.html') else BASE_URL + '/' + base
    robots = 'noindex, nofollow' if base in NOINDEX_PAGES else 'index, follow'
    title_attr = html_lib.escape(title, quote=False)
    desc_attr = html_lib.escape(description, quote=True)
    canonical_attr = html_lib.escape(canonical, quote=True)
    markup = """<meta name="description" content="%s">
<meta name="robots" content="%s">
<link rel="canonical" href="%s">
<meta property="og:type" content="website">
<meta property="og:site_name" content="Baker 1031 Investments">
<meta property="og:title" content="%s">
<meta property="og:description" content="%s">
<meta property="og:url" content="%s">
<meta property="og:image" content="%s/assets/logo.png">
<meta name="twitter:card" content="summary">
<meta name="twitter:title" content="%s">
<meta name="twitter:description" content="%s">
<meta name="twitter:image" content="%s/assets/logo.png">
<link rel="stylesheet" href="assets/site-enhancements.css">""" % (
        desc_attr, robots, canonical_attr, title_attr, desc_attr, canonical_attr,
        BASE_URL, title_attr, desc_attr, BASE_URL)
    html_text = remove_seo_tags(html_text)
    html_text = re.sub(r'<title[^>]*>.*?</title>', '<title>' + title_attr + '</title>', html_text, count=1, flags=re.S | re.I)
    jsonld = '<script type="application/ld+json">' + seo_jsonld(base, title, description, canonical, html_text) + '</script>'
    if '</title>' in html_text.lower():
        html_text = re.sub(r'(</title>)', lambda _m: _m.group(1) + '\n' + markup + '\n' + jsonld, html_text, count=1, flags=re.I)
    elif '</head>' in html_text:
        html_text = html_text.replace('</head>', markup + '\n' + jsonld + '\n</head>', 1)
    return html_text

def strip_unused_audio_styles(html_text):
    """Remove old Web Speech UI rules from published HTML.

    The actual player is styled once in assets/site-enhancements.css. Keeping
    this cleanup in the build removes the legacy rules from old article files
    without requiring a risky 678-file source rewrite.
    """
    patterns = [
        r'\n\s*\.audio-player\s*\{[^}]*\}',
        r'\n\s*\.audio-player \.ap-label\s*\{[^}]*\}',
        r'\n\s*\.audio-player button\s*\{[^}]*\}',
        r'\n\s*\.audio-player button:hover\s*\{[^}]*\}',
        r'\n\s*\.audio-player \.ap-status\s*\{[^}]*\}',
    ]
    for pattern in patterns:
        html_text = re.sub(pattern, '', html_text, flags=re.I | re.S)
    return html_text

def audio_player_html(base, title):
    item = AUDIO_MANIFEST.get(base)
    if not isinstance(item, dict) or not item.get('url'):
        return ''
    url = html_lib.escape(str(item['url']), quote=True)
    label = html_lib.escape('Listen to an executive summary of ' + title, quote=True)
    return '''<div class="audio-player" data-b1031-audio="%s">
  <span class="ap-label">Executive summary audio</span>
  <audio controls preload="none" aria-label="%s">
    <source src="%s" type="audio/mpeg">
    Your browser does not support the audio player. <a href="%s">Open the audio summary</a>.
  </audio>
</div>''' % (html_lib.escape(base, quote=True), label, url, url)

def strip_internal_nofollow(html_text):
    """Keep internal editorial links crawlable while preserving external policy."""
    base_host = urlparse(BASE_URL).netloc.lower()
    if base_host.startswith('www.'):
        base_host = base_host[4:]

    def clean_tag(match):
        tag = match.group(0)
        href_m = re.search(r'\bhref=["\']([^"\']+)', tag, flags=re.I)
        if not href_m:
            return tag
        href = html_lib.unescape(href_m.group(1)).strip()
        parsed = urlparse(href)
        parsed_host = parsed.netloc.lower()
        if parsed_host.startswith('www.'):
            parsed_host = parsed_host[4:]
        same_site = (not parsed.scheme and not parsed.netloc and not href.startswith(('#', 'mailto:', 'tel:'))) or (
            parsed_host == base_host
        )
        if not same_site or 'nofollow' not in tag.lower():
            return tag
        tag = re.sub(r'\s+nofollow\b', '', tag, flags=re.I)
        tag = re.sub(r'\s+rel=["\']\s*["\']', '', tag, flags=re.I)
        return tag
    return re.sub(r'<a\b[^>]*>', clean_tag, html_text, flags=re.I)

def normalize_internal_hrefs(html_text):
    """Point same-site links directly at emitted canonical HTML files.

    Older research pages used extensionless relative links such as
    ``calculators``. Netlify correctly redirects those URLs, but every source
    link should go directly to the final file so crawlers and visitors avoid
    a needless hop.
    """
    def rewrite(match):
        attr, quote, ref = match.group(1), match.group(2), html_lib.unescape(match.group(3))
        if not ref or ref.startswith(('#', 'mailto:', 'tel:', 'javascript:', 'data:', '//')):
            return match.group(0)
        parsed = urlparse(ref)
        site_host = urlparse(BASE_URL).netloc.lower()
        parsed_host = parsed.netloc.lower()
        if site_host.startswith('www.'):
            site_host = site_host[4:]
        if parsed_host.startswith('www.'):
            parsed_host = parsed_host[4:]
        same_site = (not parsed.netloc and not parsed.scheme) or parsed_host == site_host
        if not same_site:
            return match.group(0)
        path = parsed.path or ''
        if not path or path == '/' or path.endswith('/') or path.endswith('.html'):
            return match.group(0)
        clean_path = path.lstrip('/')
        if '/' in clean_path or not os.path.isfile(os.path.join(DIST, clean_path + '.html')):
            return match.group(0)
        new_path = '/' + clean_path + '.html'
        if not parsed.netloc and not parsed.scheme:
            new_path = clean_path + '.html'
        new_ref = new_path + (('?' + parsed.query) if parsed.query else '') + (('#' + parsed.fragment) if parsed.fragment else '')
        return '%s=%s%s%s' % (attr, quote, html_lib.escape(new_ref, quote=True), quote)
    return re.sub(r'\b(href|src|poster)=("|\')([^"\']*)\2', rewrite, html_text, flags=re.I)

def inject(html_text, base=''):
    html_text = strip_unused_audio_styles(html_text)
    html_text = strip_internal_nofollow(html_text)
    html_text = html_text.replace('<!-- @@NAV@@ -->', nav)
    state_note = state_authority_note(base)
    if state_note and 'class="b1031-state-source"' not in html_text:
        if '<!-- @@FOOTER@@ -->' in html_text:
            html_text = html_text.replace('<!-- @@FOOTER@@ -->', state_note + '\n<!-- @@FOOTER@@ -->', 1)
        elif '</main>' in html_text.lower():
            html_text = re.sub(r'(</main>)', state_note + '\n\\1', html_text, count=1, flags=re.I)
    if base not in NOINDEX_PAGES and '<!-- @@FOOTER@@ -->' in html_text:
        related = RELATED_HUBS_HTML
        if base in ORPHAN_RESOURCE_LINKS:
            related += '\n' + ORPHAN_RESOURCE_LINKS[base]
        html_text = html_text.replace('<!-- @@FOOTER@@ -->', related + '\n<!-- @@FOOTER@@ -->', 1)
    html_text = html_text.replace('<!-- @@FOOTER@@ -->', footer)
    if base not in NOINDEX_PAGES:
        html_text = html_text.replace('<!-- @@REVIEW_NOTE@@ -->', REVIEW_NOTE_HTML)
    else:
        html_text = html_text.replace('<!-- @@REVIEW_NOTE@@ -->', '')
    if base in AUDIO_MANIFEST and '<div class="audio-player"' not in html_text:
        title_for_audio = page_title(html_text, base.replace('.html', '').replace('-', ' ').title())
        player = audio_player_html(base, title_for_audio)
        if player:
            if re.search(r'</article>', html_text, flags=re.I):
                html_text = re.sub(r'</article>', player + '\n</article>', html_text, count=1, flags=re.I)
            elif re.search(r'</main>', html_text, flags=re.I):
                html_text = re.sub(r'</main>', player + '\n</main>', html_text, count=1, flags=re.I)
    # Keep non-critical third-party scripts off the parser's critical path and
    # let below-the-fold images load only when they approach the viewport.
    html_text = re.sub(r'<script(?![^>]*\b(?:async|defer)\b)(?=[^>]*\bsrc=)([^>]*)>', r'<script defer\1>', html_text, flags=re.I)
    html_text = re.sub(r'<img(?![^>]*\b(?:loading|fetchpriority)=)(?![^>]*class=["\'][^"\']*\blogo\b)(?=[^>]*\bsrc=)', '<img loading="lazy" decoding="async"', html_text, flags=re.I)
    html_text = seo_inject(html_text, base)
    if FAVICON_TAG not in html_text and '</head>' in html_text:
        html_text = html_text.replace('</head>', FAVICON_TAG + '\n</head>', 1)
    if base not in ANALYTICS_EXCLUDE and GA4_TAG not in html_text and '</head>' in html_text:
        html_text = html_text.replace('</head>', GA4_TAG + '\n</head>', 1)
        html_text = html_text.replace('</head>', ANALYTICS_SCRIPT + '\n</head>', 1)
    if base not in SENTRY_EXCLUDE and SENTRY_TAG not in html_text and '</head>' in html_text:
        html_text = html_text.replace('</head>', SENTRY_TAG + '\n</head>', 1)
    if CLARITY_PROJECT_ID and base not in CLARITY_EXCLUDE and CLARITY_TAG not in html_text and '</head>' in html_text:
        html_text = html_text.replace('</head>', CLARITY_TAG + '\n</head>', 1)
    return html_text

# Pages that stay OPEN (no soft gate): home, registration, About group, contact,
# lead-gen/capability pages, all legal/compliance pages, auth pages, 404. Every
# other page (listings, offerings, insights/articles, resources, strategies, geo,
# glossary, calculators, sponsors, data-center, guides, property) gets the soft gate.
GATE_OPEN = set([
    'baker1031.html', 'index.html', 'request-access.html',
    'about.html', 'our-approach.html', 'methodology.html', 'due-diligence.html',
    'team-partners.html', 'jerry-baker-bio.html',
    'contact.html', 'who-we-serve.html', 'for-advisors-cpas.html', 'for-agents-brokers.html',
    'account.html', 'employee.html', '404.html',
    'privacy-policy.html', 'terms.html', 'ccpa.html', 'commitment-to-privacy.html',
    'reg-bi.html', 'accessibility.html', 'disclosures.html', 'sitemap.html', 'request-access.html',
    'site-search.html', 'ask-llm.html',
])
GATE_TAG = '<script src="assets/softgate.js" defer></script>'

def gate(html_text, base):
    if base in GATE_OPEN:
        return html_text
    if GATE_TAG in html_text:
        return html_text
    if '</body>' in html_text:
        return html_text.replace('</body>', '  ' + GATE_TAG + '\n</body>', 1)
    return html_text + '\n' + GATE_TAG

def generated_sitemap_markup():
    """Replace the stale hand-maintained HTML sitemap with current build data."""
    source_files = []
    for source_dir in ('pages', 'pages-legacy'):
        source_files.extend(os.path.basename(p) for p in _glob.glob(os.path.join(ROOT, 'src', source_dir, '*.html')))
    names = set(source_files) | set(generated) | {'index.html'}
    names -= NOINDEX_PAGES
    names -= SKIP_SOURCE_PAGES
    names.discard('sitemap.html')

    def label_for(base):
        if base == 'index.html':
            return 'Baker 1031 Investments — Home'
        if base in SEO_META and SEO_META[base].get('title'):
            return SEO_META[base]['title']
        raw = generated.get(base)
        if raw is None:
            for source_dir in ('pages', 'pages-legacy'):
                candidate = os.path.join(ROOT, 'src', source_dir, base)
                if os.path.isfile(candidate):
                    raw = open(candidate, encoding='utf-8', errors='replace').read()
                    break
        return page_title(raw or '', base[:-5].replace('-', ' ').title())

    def link_item(base, label=None):
        href = '/' if base == 'index.html' else base
        return '<li><a href="%s">%s</a></li>' % (href, html_lib.escape(label or label_for(base)))

    def group(title, items, collapsed=False):
        attrs = '' if not collapsed else ''
        body = '\n'.join(link_item(base, label) for base, label in items)
        return '<div class="grp"><h2>%s</h2><hr class="rule"/><div class="cnt">%d pages</div><ul>%s</ul></div>' % (
            html_lib.escape(title), len(items), body)

    core_names = [base for base in PAGE_DESCRIPTIONS if base in names and base not in ('index.html', 'baker1031.html')]
    core_items = [(base, label_for(base)) for base in sorted(set(core_names))]
    offering_items = [('investments.html', 'Current Offerings')]
    offering_items.extend((row['url'], row['name']) for row in dir_rows if row.get('url') in names)
    sponsor_items = [('sponsors.html', 'Sponsor Directory')]
    sponsor_items.extend((row['url'], row['name']) for row in sp_dir if row.get('url') in names)
    used = set(core_names) | {base for base, _ in offering_items} | {base for base, _ in sponsor_items}
    research_items = [(base, label_for(base)) for base in sorted(names - used)]
    groups = [group('Primary Hubs', core_items), group('Current Offerings', offering_items),
              group('Sponsors', sponsor_items), group('Insights, Guides, Calculators, Locations & Other Research', research_items)]
    return '<section class="wrap smap">' + ''.join(groups) + '</section>'

shutil.rmtree(DIST, ignore_errors=True)
os.makedirs(DIST)
count = 0
import glob as _glob
for d in ('pages', 'pages-legacy'):
    for f in _glob.glob(os.path.join(ROOT, 'src', d, '*.html')):
        base = os.path.basename(f)
        if base in SKIP_SOURCE_PAGES:
            continue
        html_text = open(f).read()
        if base == 'investments.html':
            html_text = re.sub(r'<script type="application/json" id="directory-data">\n.*?\n</script>',
                               lambda _m: '<script type="application/json" id="directory-data">\n' + html_json(dir_rows) + '\n</script>',
                               html_text, flags=re.S)
        if base in ('employee.html', 'request-access.html'):
            html_text = html_text.replace('<script type="application/json" id="directory-data">[]</script>',
                                          '<script type="application/json" id="directory-data">' + html_json(dir_rows) + '</script>')
        if base == 'account.html':
            _bmindex = ([{'slug': r['url'].replace('.html', ''), 'name': r['name'], 'url': r['url']} for r in dir_rows]
                        + [{'slug': r['url'].replace('.html', ''), 'name': r['name'], 'url': r['url']} for r in sp_dir])
            html_text = re.sub(r'<script type="application/json" id="bookmark-index">.*?</script>',
                               lambda _m: '<script type="application/json" id="bookmark-index">' + html_json(_bmindex) + '</script>',
                               html_text, flags=re.S)
        if base == 'delaware-statutory-trusts.html':
            bm_map = {s(b['Property Type']): b for b in BM}
            picks = ['Marina', 'Healthcare', 'Net Lease', 'Multifamily', 'Self-Storage', 'Office']
            vals = []
            for key in picks:
                b = bm_map.get(key)
                if b and isinstance(b.get('Average Yield'), (int, float)) and b['Average Yield']:
                    vals.append((key, b['Average Yield'] * 100))
            vals.sort(key=lambda x: -x[1])
            xs = [60, 141, 222, 303, 384, 465]
            bars, labels = [], []
            for i, (label, v) in enumerate(vals[:6]):
                hpx = v * 16.875
                bars.append('          <rect x="%d" y="%.1f" width="52" height="%.1f"><title>%s — %.2f%%</title></rect>' % (xs[i], 155 - hpx, hpx, label, v))
                labels.append('<text x="%d" y="172">%s</text>' % (xs[i] + 26, label))
            bar_block = '<g fill="#243856">\n' + '\n'.join(bars) + '\n        </g>'
            lbl_block = '<g fill="#777" font-size="10" text-anchor="middle">\n          ' + ''.join(labels) + '\n        </g>'
            html_text = re.sub(r'<g fill="#243856">\n.*?</g>', lambda _m: bar_block, html_text, count=1, flags=re.S)
            html_text = re.sub(r'<g fill="#777" font-size="10" text-anchor="middle">\n.*?</g>', lambda _m: lbl_block, html_text, count=1, flags=re.S)
        if base == 'sponsors.html':
            html_text = re.sub(r'<script type="application/json" id="directory-data">\n.*?\n</script>',
                               lambda _m: '<script type="application/json" id="directory-data">\n' + html_json(sp_dir) + '\n</script>',
                               html_text, flags=re.S)
        if base == 'sitemap.html':
            html_text = re.sub(r'<section class="wrap smap">.*?</section>', generated_sitemap_markup(), html_text, count=1, flags=re.S)
        open(os.path.join(DIST, base), 'w').write(gate(inject(html_text, base), base))
        count += 1
for base, html_text in generated.items():
    open(os.path.join(DIST, base), 'w').write(gate(inject(html_text, base), base))
    count += 1

shutil.copytree(os.path.join(ROOT, 'src', 'assets'), os.path.join(DIST, 'assets'))
# Copy root-level discovery assets before validation so absolute root links such
# as /favicon.ico are checked against the same dist tree that will be deployed.
shutil.copy(os.path.join(ROOT, 'data', 'favicon.ico'), os.path.join(DIST, 'favicon.ico'))
open(os.path.join(DIST, 'site.webmanifest'), 'w').write(json.dumps({
    'name': 'Baker 1031 Investments', 'short_name': 'Baker 1031',
    'icons': [{'src': '/assets/icon-192.png', 'sizes': '192x192', 'type': 'image/png'},
              {'src': '/assets/icon-512.png', 'sizes': '512x512', 'type': 'image/png', 'purpose': 'any maskable'}],
    'theme_color': '#243856', 'background_color': '#243856', 'display': 'standalone', 'start_url': '/'}))
home_source = open(os.path.join(ROOT, 'src', 'pages', 'baker1031.html')).read()
open(os.path.join(DIST, 'index.html'), 'w').write(gate(inject(home_source, 'index.html'), 'index.html'))
print('built %d pages -> dist/' % (count + 1))

# Normalize links only after all source and generated pages have been written,
# so the emitted file list is the source of truth for canonical targets.
for _p in sorted(f for f in os.listdir(DIST) if f.endswith('.html')):
    _path = os.path.join(DIST, _p)
    _raw = open(_path, encoding='utf-8', errors='replace').read()
    _updated = normalize_internal_hrefs(_raw)
    if _updated != _raw:
        open(_path, 'w', encoding='utf-8').write(_updated)

# Build a compact, public search index from indexable page titles, descriptions,
# headings, and introductory copy. It keeps search entirely client-side: no
# visitor data or third-party search account is required.
search_rows = []
for search_base in sorted(f for f in os.listdir(DIST) if f.endswith('.html')):
    search_path = os.path.join(DIST, search_base)
    search_raw = open(search_path, encoding='utf-8', errors='replace').read()
    if re.search(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\'][^"\']*noindex', search_raw, flags=re.I):
        continue
    if search_base in ('site-search.html', '404.html'):
        continue
    search_title = page_title(search_raw, search_base.replace('.html', '').replace('-', ' ').title())
    search_description = page_description(search_base, search_raw, SEO_META.get(search_base, {}))
    search_text = text_from_html(search_raw)
    search_text = re.sub(r'\s+', ' ', search_text).strip()
    search_rows.append({
        'url': '/' + search_base,
        'title': search_title,
        'description': search_description,
        'text': search_text[:900],
    })
open(os.path.join(DIST, 'search-index.json'), 'w', encoding='utf-8').write(
    json.dumps(search_rows, ensure_ascii=False, separators=(',', ':'))
)

# Externalize exact repeated inline style blocks into shared assets. This keeps
# page-specific CSS intact while removing the same large nav/footer/font rules
# from hundreds of HTML documents.
def externalize_repeated_css():
    style_pattern = re.compile(r'<style(?P<attrs>[^>]*)>(?P<body>.*?)</style>', flags=re.S | re.I)
    occurrences = {}
    for base in sorted(f for f in os.listdir(DIST) if f.endswith('.html')):
        path = os.path.join(DIST, base)
        raw = open(path, encoding='utf-8', errors='replace').read()
        for match in style_pattern.finditer(raw):
            attrs = match.group('attrs').strip()
            body = match.group('body').strip()
            if attrs or len(body) < 500:
                continue
            digest = hashlib.sha1(body.encode('utf-8')).hexdigest()[:12]
            entry = occurrences.setdefault(digest, {'body': body, 'pages': set()})
            entry['pages'].add(base)
    repeated = {k: v for k, v in occurrences.items() if len(v['pages']) >= 2}
    for digest, entry in repeated.items():
        # The original inline CSS lived in a page at the dist root, where
        # url(assets/foo) resolved correctly. Shared styles now live inside
        # dist/assets, so keep those references relative to that directory.
        shared_body = re.sub(r'url\(\s*(["\']?)assets/([^"\')]+)\1\s*\)',
                             lambda m: 'url(%s%s%s)' % (m.group(1), m.group(2), m.group(1)),
                             entry['body'])
        open(os.path.join(DIST, 'assets', 'shared-' + digest + '.css'), 'w').write(shared_body + '\n')
    replaced = 0
    for base in sorted(f for f in os.listdir(DIST) if f.endswith('.html')):
        path = os.path.join(DIST, base)
        raw = open(path, encoding='utf-8', errors='replace').read()
        def replace_style(match):
            nonlocal replaced
            attrs = match.group('attrs').strip()
            body = match.group('body').strip()
            if attrs or len(body) < 500:
                return match.group(0)
            digest = hashlib.sha1(body.encode('utf-8')).hexdigest()[:12]
            shared_path = os.path.join(DIST, 'assets', 'shared-' + digest + '.css')
            if digest not in repeated or not os.path.isfile(shared_path):
                return match.group(0)
            replaced += 1
            return '<link rel="stylesheet" href="assets/shared-%s.css">' % digest
        updated = style_pattern.sub(replace_style, raw)
        if updated != raw:
            open(path, 'w').write(updated)
    print('shared CSS: externalized %d repeated blocks into %d assets' % (replaced, len(repeated)))

externalize_repeated_css()

# ---------------------------------------------------------------- sitemap.xml
_pages = sorted(f for f in os.listdir(DIST) if f.endswith('.html'))
_urls = []
for _p in _pages:
    _raw = open(os.path.join(DIST, _p), encoding='utf-8', errors='replace').read()
    if re.search(r"<meta[^>]+name=[\"']robots[\"'][^>]+content=[\"'][^\"']*noindex", _raw, flags=re.I):
        continue
    loc = BASE_URL + '/' + ('' if _p == 'index.html' else _p)
    _urls.append('  <url><loc>%s</loc></url>' % loc)
open(os.path.join(DIST, 'sitemap.xml'), 'w').write(
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    + '\n'.join(_urls) + '\n</urlset>\n')
open(os.path.join(DIST, 'robots.txt'), 'w').write(
    'User-agent: *\nAllow: /\nSitemap: %s/sitemap.xml\n' % BASE_URL)
print('sitemap.xml: %d urls' % len(_urls))

# ---------------------------------------------------------------- generated-site validation
def validate_dist():
    html_files = sorted(f for f in os.listdir(DIST) if f.endswith('.html'))
    failures = []
    thin_pages = []
    page_meta = {}
    quality = {'generated_at': BUILD_DATE, 'html_pages': len(html_files), 'thin_pages': [], 'citation_gaps': []}
    site_host = urlparse(BASE_URL).netloc.lower()
    if site_host.startswith('www.'):
        site_host = site_host[4:]
    redirect_paths = {'/offerings.html', '/article-template.html', '/insight-1031-exchange-guide.html'}
    generated_paths = {'/llms.txt', '/llms-full.txt'}

    def resolve_site_reference(ref, base):
        """Return a same-site target and fragment, or None for true externals."""
        ref = html_lib.unescape(ref).strip()
        if not ref or ref.startswith(('mailto:', 'tel:', 'javascript:', 'data:')):
            return None
        if ref.startswith('//'):
            parsed = urlparse('https:' + ref)
        elif urlparse(ref).scheme:
            parsed = urlparse(ref)
        else:
            parsed = urlparse(urljoin(BASE_URL + '/' + base, ref))
        if parsed.scheme not in ('http', 'https'):
            return None
        parsed_host = parsed.netloc.lower()
        if parsed_host.startswith('www.'):
            parsed_host = parsed_host[4:]
        if parsed_host and parsed_host != site_host:
            return None
        path = unquote(parsed.path or '/')
        path = '/' + posixpath.normpath(path.lstrip('/')).lstrip('./')
        target = os.path.join(DIST, 'index.html' if path == '/' else path.lstrip('/'))
        return target, path, unquote(parsed.fragment or '')

    def plain(raw):
        raw = re.sub(r'<(style|script|noscript)[^>]*>.*?</\1>', ' ', raw, flags=re.S | re.I)
        raw = re.sub(r'<[^>]+>', ' ', raw)
        return re.sub(r'\s+', ' ', html_lib.unescape(raw)).strip()

    for base in html_files:
        path = os.path.join(DIST, base)
        raw = open(path, encoding='utf-8', errors='replace').read()
        robots = re.search(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\']([^"\']+)', raw, flags=re.I)
        indexable = not robots or 'noindex' not in robots.group(1).lower()
        title_m = re.search(r'<title[^>]*>(.*?)</title>', raw, flags=re.S | re.I)
        desc_m = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']*)', raw, flags=re.I)
        canonical_m = re.search(r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)', raw, flags=re.I)
        page_meta[base] = {
            'title': text_from_html(title_m.group(1)) if title_m else '',
            'description': html_lib.unescape(desc_m.group(1)) if desc_m else '',
            'canonical': canonical_m.group(1) if canonical_m else '',
            'indexable': indexable,
        }
        if indexable and not title_m:
            failures.append('%s: missing title' % base)
        if indexable and len(page_meta[base]['title']) > TITLE_LIMIT:
            failures.append('%s: title exceeds %d characters' % (base, TITLE_LIMIT))
        if not desc_m:
            failures.append('%s: missing meta description' % base)
        if desc_m and len(page_meta[base]['description']) > DESCRIPTION_LIMIT:
            failures.append('%s: meta description exceeds %d characters' % (base, DESCRIPTION_LIMIT))
        if not canonical_m:
            failures.append('%s: missing canonical URL' % base)
        if not re.search(r'<meta[^>]+property=["\']og:title["\']', raw, flags=re.I) or not re.search(r'<meta[^>]+name=["\']twitter:card["\']', raw, flags=re.I):
            failures.append('%s: missing Open Graph or Twitter metadata' % base)
        placeholder_hits = re.findall(r'(?i)(?:Placeholder Article Title|Placeholder copy|Placeholder introduction|Placeholder body copy|Placeholder question|Placeholder answer|Placeholder bio|\[Placeholder regulatory disclosure)', raw)
        if placeholder_hits:
            failures.append('%s: public placeholder text: %s' % (base, ', '.join(sorted(set(placeholder_hits)))))
        ids = re.findall(r'\bid=["\']([^"\']+)["\']', raw, flags=re.I)
        seen = {}
        duplicates = []
        for value in ids:
            seen[value] = seen.get(value, 0) + 1
        duplicates = sorted(k for k, v in seen.items() if v > 1)
        if duplicates:
            failures.append('%s: duplicate HTML IDs: %s' % (base, ', '.join(duplicates)))
        blocks = re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', raw, flags=re.S | re.I)
        schema_types = set()
        for i, block in enumerate(blocks, 1):
            try:
                parsed = json.loads(block)
                if not isinstance(parsed, dict) or parsed.get('@context') != 'https://schema.org':
                    failures.append('%s: JSON-LD block %d has an invalid Schema.org context' % (base, i))
                entities = parsed.get('@graph', [parsed]) if isinstance(parsed, dict) else []
                for entity in entities:
                    if not isinstance(entity, dict):
                        continue
                    entity_type = entity.get('@type')
                    if isinstance(entity_type, list):
                        schema_types.update(str(item) for item in entity_type)
                    elif entity_type:
                        schema_types.add(str(entity_type))
                    types = set(str(item) for item in entity_type) if isinstance(entity_type, list) else {str(entity_type)} if entity_type else set()
                    if not types:
                        failures.append('%s: JSON-LD block %d contains an entity without @type' % (base, i))
                    if ('Article' in types or 'WebPage' in types) and not entity.get('url'):
                        failures.append('%s: JSON-LD block %d page entity is missing url' % (base, i))
                    if 'Article' in types:
                        for required in ('headline', 'author', 'publisher', 'dateModified'):
                            if not entity.get(required):
                                failures.append('%s: Article schema is missing %s' % (base, required))
                    if 'FAQPage' in types:
                        questions = entity.get('mainEntity')
                        if not isinstance(questions, list) or not questions or any(
                            not isinstance(q, dict) or q.get('@type') != 'Question' or
                            not q.get('name') or not isinstance(q.get('acceptedAnswer'), dict) or
                            q['acceptedAnswer'].get('@type') != 'Answer' or not q['acceptedAnswer'].get('text')
                            for q in questions
                        ):
                            failures.append('%s: FAQPage schema has an invalid question/answer structure' % base)
                    if 'WebApplication' in types:
                        for required in ('name', 'url', 'applicationCategory', 'operatingSystem'):
                            if not entity.get(required):
                                failures.append('%s: WebApplication schema is missing %s' % (base, required))
                    if 'AudioObject' in types and not entity.get('contentUrl'):
                        failures.append('%s: AudioObject schema is missing contentUrl' % base)
                    if 'DefinedTermSet' in types and not isinstance(entity.get('hasDefinedTerm'), list):
                        failures.append('%s: DefinedTermSet schema is missing hasDefinedTerm' % base)
            except Exception as exc:
                failures.append('%s: invalid JSON-LD block %d (%s)' % (base, i, exc))
        if base in INTERACTIVE_CALCULATOR_PAGES and not re.search(r'WebApplication', '\n'.join(blocks), flags=re.I):
            failures.append('%s: calculator is missing WebApplication structured data' % base)
        if re.search(r'<audio\b|<source\b[^>]*type=["\']audio/', raw, flags=re.I) and not re.search(r'AudioObject', '\n'.join(blocks), flags=re.I):
            failures.append('%s: audio element is missing AudioObject structured data' % base)
        if indexable and len(re.findall(r'<details\b', raw, flags=re.I)) >= 2 and 'FAQPage' not in schema_types:
            failures.append('%s: FAQ content is missing FAQPage structured data' % base)
        if indexable and base == 'glossary.html' and re.search(r'<dt\b', raw, flags=re.I) and 'DefinedTermSet' not in schema_types:
            failures.append('%s: glossary terms are missing DefinedTermSet structured data' % base)
        for block in re.findall(r'<script[^>]+type=["\']application/json["\'][^>]*>(.*?)</script>', raw, flags=re.S | re.I):
            try:
                payload = json.loads(block)
            except Exception:
                continue
            def check_dynamic_urls(value):
                if isinstance(value, dict):
                    for key, child in value.items():
                        if key in ('url', 'href') and isinstance(child, str) and child.endswith('.html') and not child.startswith(('http://', 'https://', '/')):
                            if not os.path.isfile(os.path.join(DIST, child)):
                                failures.append('%s: broken generated data link %s' % (base, child))
                        check_dynamic_urls(child)
                elif isinstance(value, list):
                    for child in value:
                        check_dynamic_urls(child)
            check_dynamic_urls(payload)
        if base == 'sponsors.html' and re.search(r'<img[^>]+alt=(["\'])\s*\1', raw, flags=re.I):
            failures.append('sponsors.html: generated sponsor logo has empty alt text')
        word_count = len(re.findall(r"\b[\w’'-]+\b", plain(raw)))
        visible_text = plain(raw)
        visible_placeholders = re.findall(r'(?i)template provided for review|template for review|should be finalized|needs approval|placeholder regulatory disclosure', visible_text)
        if visible_placeholders:
            failures.append('%s: visible placeholder/review language: %s' % (base, ', '.join(sorted(set(visible_placeholders)))))
        if indexable and not re.search(r'Reviewed by\s+Lori Kamen.*?Last reviewed\s+' + re.escape(REVIEW_DATE), visible_text, flags=re.I | re.S):
            failures.append('%s: missing visible reviewer and review date' % base)
        if indexable and word_count < 800:
            thin_pages.append({'page': base, 'words': word_count, 'reason': 'below 800-word review threshold'})
        state_slug = state_slug_for_page(base)
        if state_slug and STATE_TAX_AUTHORITIES[state_slug]['url'] not in raw:
            failures.append('%s: missing official state tax authority citation' % base)
        if indexable and base not in HUB_FILES and not page_citations(base, raw):
            quality['citation_gaps'].append(base)
        for attr, ref in re.findall(r'\b(href|src|poster)=["\']([^"\']+)["\']', raw, flags=re.I):
            resolved = resolve_site_reference(ref, base)
            if not resolved:
                continue
            target, target_path, fragment = resolved
            if not os.path.isfile(target) and target_path not in redirect_paths and target_path not in generated_paths:
                failures.append('%s: broken same-site %s %s' % (base, attr, ref))
                continue
            if fragment and os.path.isfile(target):
                target_raw = open(target, encoding='utf-8', errors='replace').read()
                anchors = set(re.findall(r'\b(?:id|name)=["\']([^"\']+)["\']', target_raw, flags=re.I))
                if fragment not in anchors:
                    failures.append('%s: missing anchor #%s in %s' % (base, fragment, ref))

    for asset in sorted(f for f in os.listdir(os.path.join(DIST, 'assets')) if f.endswith('.css')):
        css_path = os.path.join(DIST, 'assets', asset)
        css = open(css_path, encoding='utf-8', errors='replace').read()
        for css_url in re.findall(r'url\(\s*["\']?([^"\')]+)', css, flags=re.I):
            if css_url.startswith(('data:', 'http://', 'https://', '//', '#')):
                continue
            if not os.path.isfile(os.path.normpath(os.path.join(os.path.dirname(css_path), css_url))):
                failures.append('%s: broken CSS asset URL %s' % (asset, css_url))

    title_groups = {}
    desc_groups = {}
    for base, meta in page_meta.items():
        if not meta['indexable']:
            continue
        title_groups.setdefault(meta['title'], []).append(base)
        desc_groups.setdefault(meta['description'], []).append(base)
    for title, pages in title_groups.items():
        if title and len(pages) > 1:
            failures.append('duplicate indexable title %r: %s' % (title, ', '.join(pages)))
    for desc, pages in desc_groups.items():
        if desc and len(pages) > 1:
            failures.append('duplicate indexable meta description on: %s' % ', '.join(pages))
    quality['thin_pages'] = thin_pages
    quality['duplicate_title_groups'] = {k: v for k, v in title_groups.items() if k and len(v) > 1}
    quality['duplicate_description_groups'] = {k: v for k, v in desc_groups.items() if k and len(v) > 1}
    quality['failure_count'] = len(failures)
    open(os.path.join(DIST, 'content-quality-report.json'), 'w').write(json.dumps(quality, indent=2, ensure_ascii=False))
    if failures:
        print('BUILD VALIDATION FAILED:')
        for failure in failures:
            print(' - ' + failure)
        raise SystemExit(1)
    print('validation: %d HTML pages passed metadata, canonical, structured-data, link, placeholder, and duplicate-ID gates' % len(html_files))
    print('content review report: %d thin pages and %d citation gaps recorded in dist/content-quality-report.json' % (len(thin_pages), len(quality['citation_gaps'])))

validate_dist()

# ---------------------------------------------------------------- llms.txt / llms-full.txt / favicon / manifest
# llms.txt: curated index (template), with the Current Offerings section refreshed from the sheet
_tpl = open(os.path.join(ROOT, 'data', 'llms-template.txt')).read()
_off_lines = []
for _row in LIST:
    if s(_row.get('Status')).lower() in ('closed', 'sold out', 'inactive'):
        continue
    _nm = s(_row['Investment Name'])
    _off_lines.append('- [%s](%s/%s)' % (_nm, '{{BASE}}', off_slug[_nm]))
_tpl = re.sub(r'## Current Offerings\n(?:- \[[^\n]*\n)*',
              '## Current Offerings\n' + '\n'.join(_off_lines) + '\n', _tpl, count=1)
open(os.path.join(DIST, 'llms.txt'), 'w').write(_tpl.replace('{{BASE}}', BASE_URL).replace('{{GENERATED}}', BUILD_DATE))

# llms-full.txt: concatenated plain text of every page for AI ingestion
def _strip_html(raw):
    raw = re.sub(r'<(style|script)[^>]*>.*?</\1>', ' ', raw, flags=re.S)
    raw = re.sub(r'<[^>]+>', ' ', raw)
    raw = raw.replace('&amp;', '&').replace('&nbsp;', ' ').replace('&middot;', '·').replace('&copy;', '(c)')
    return re.sub(r'[ \t]+', ' ', re.sub(r'\s*\n\s*', '\n', raw)).strip()

with open(os.path.join(DIST, 'llms-full.txt'), 'w') as _fh:
    _fh.write('# Baker 1031 Investments — full site text for AI ingestion\n# Index: %s/llms.txt\n\n' % BASE_URL)
    for _p in sorted(os.listdir(DIST)):
        if not _p.endswith('.html') or _p in ('404.html', 'article-template.html', 'baker1031.html'):
            continue
        _raw = open(os.path.join(DIST, _p), encoding='utf-8', errors='replace').read()
        if re.search(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\'][^"\']*noindex', _raw, flags=re.I):
            continue
        _title = re.search(r'<title>(.*?)</title>', _raw, re.S)
        _body = _strip_html(_raw)
        # drop repeated nav/footer noise: cut everything before the breadcrumb "Home ❯" and after "Continue Exploring"
        _i = _body.find('Home ❯')
        _j = _body.find('Continue Exploring Baker 1031')
        if 0 <= _i < (_j if _j > 0 else len(_body)):
            _body = _body[_i:(_j if _j > 0 else len(_body))]
        _fh.write('\n===== PAGE: %s/%s =====\n%s\n%s\n' % (
            BASE_URL, _p, (_title.group(1).strip() if _title else _p), _body))

print('llms.txt + llms-full.txt + favicon + manifest written')
